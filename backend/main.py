import csv
import io
import json
import math
import os
import re
import subprocess
import glob
import time
import threading
import zipfile
from datetime import date, datetime, timedelta, timezone
from typing import Optional, Tuple

import httpx
import pandas as pd
from dotenv import load_dotenv
from fastapi import Body, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import text
from watchdog.observers.polling import PollingObserver
from watchdog.events import FileSystemEventHandler

from excel_mapper import ExcelParseError, parse_excel, validate_excel
from forecast_engine import (
    confidence_bounds,
    forecast_with_method,
    holdout_accuracy,
)
from database import (
    DB_TYPE,
    DATABASE_URL,
    PG_FALLBACK_REASON,
    ForecastMeta,
    Item,
    MRPResult,
    OpenPO,
    SalesHistory,
    SessionLocal,
    SQL_VIEWS,
    VIEW_CREATE_SQL,
    WEEKS,
    create_sql_views,
    get_db_display_url,
    is_readonly_query,
)

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WATCHED_DIR = os.path.join(BASE_DIR, "watched_files")
os.makedirs(WATCHED_DIR, exist_ok=True)

sync_state = {
    "filename": None,
    "last_updated": None,
    "last_mtime": 0.0,
    "source": None,
}

watch_state = {
    "active": False,
    "watched_file": None,
}

# In-memory only (never persisted). False on server start; True only after an
# intentional load (upload / load-sample / live file drop) during this session.
session_state = {"active": False}
last_excel_content: Optional[bytes] = None

SAMPLE_FILE_PATH = os.path.join(BASE_DIR, "sample_mrp_data.xlsx")
MRP_DB_PATH = os.path.join(BASE_DIR, "mrp.db")

EXPORT_MRP_COLUMNS = [
    "item_id", "item_name", "week", "gross_req", "scheduled_receipts",
    "projected_inventory", "net_req", "planned_order", "safety_stock",
    "stockout_risk", "need_date", "release_date", "release_week", "is_overdue",
    "fg_production_date",
]
EXPORT_ITEMS_COLUMNS = [
    "item_id", "item_name", "lead_time_weeks", "safety_stock",
    "lot_size", "unit", "unit_cost", "available_qty", "category", "bom_level",
]
EXPORT_POS_COLUMNS = [
    "po_number", "item_id", "supplier", "order_qty", "expected_receipt_week", "status",
]

POWERBI_INSTRUCTIONS = f"""MRP Data Export for Power BI / Excel
=====================================
Files included:
- mrp_results.csv: Weekly MRP calculations (gross req, net req, projected inventory, planned orders, release dates)
- items.csv: Item master data (lead times, safety stock, lot sizes, costs)
- open_pos.csv: Open purchase orders with status and arrival weeks

To import into Microsoft Fabric (free, browser-based):
1. Go to app.fabric.microsoft.com (free Microsoft account)
2. Create a new report
3. Upload these CSV files as your data source
4. Connect the tables: mrp_results joins to items on item_id, mrp_results joins to open_pos on item_id

Suggested visuals to build:
1. Stacked bar chart: Gross Req vs Net Req vs Planned Order by week (use mrp_results, filter by item)
2. Matrix heatmap: Rows=item_name, Columns=week, Values=projected_inventory, colored by stockout_risk
3. Line chart: Projected inventory trend per item over 8 weeks
4. Table: Planned order releases sorted by release_date showing item_name, planned_order, release_date, need_date
5. KPI cards: Total stockout risks, Total planned orders this week, Total PO value

To connect Power BI Desktop directly to the live database (Windows only):
- Install SQLite ODBC driver from sqliteodbc.sourceforge.net
- Get Data → ODBC → SQLite3 ODBC Driver
- Database path: {MRP_DB_PATH}"""

_process_lock = threading.Lock()
_debounce_lock = threading.Lock()
_debounce_timer = None
_pending_path = None
_observer = None

app = FastAPI(title="MRP API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health():
    return {"status": "ok"}


class ChatRequest(BaseModel):
    message: str
    api_key: Optional[str] = None


class SqlQueryRequest(BaseModel):
    query: str


class ScenarioRequest(BaseModel):
    item_id: str
    demand_pct: float = 0
    lead_time_delta: int = 0
    safety_stock_pct: float = 0
    demand_pct_change: Optional[float] = None
    lead_time_delay_weeks: Optional[int] = None
    safety_stock_pct_change: Optional[float] = None


class ForecastParams(BaseModel):
    ma_window: int = 3
    alpha: float = 0.3
    beta: float = 0.1
    trend: str = "add"
    seasonal: str = "add"
    seasonal_periods: int = 4


class ForecastRequest(BaseModel):
    item_id: str
    method: str = "exponential_smoothing"
    periods: int = 8
    params: Optional[ForecastParams] = None


class ForecastApplyRequest(BaseModel):
    item_id: str
    forecast_values: list
    method: Optional[str] = None


class ForecastOverrideRequest(BaseModel):
    item_id: str
    week: str
    value: float


def week_index(week: str) -> int:
    return WEEKS.index(week) if week in WEEKS else -1


def get_planning_start_monday() -> date:
    today = date.today()
    return today - timedelta(days=today.weekday())


def week_to_need_date(week: str, planning_start: date) -> date:
    return planning_start + timedelta(days=week_index(week) * 7)


def format_plan_date(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def date_to_week_bucket(d: date, planning_start: date) -> str:
    delta_days = (d - planning_start).days
    if delta_days >= 0:
        return f"W{delta_days // 7 + 1}"
    weeks_past = (-delta_days - 1) // 7 + 1
    return f"W-{weeks_past}"


def compute_summary_metrics(mrp_rows, items, pos):
    item_map = {i.item_id: i for i in items}
    stockout_risks = [
        {
            "item_id": r.item_id,
            "week": r.week,
            "projected_inventory": r.projected_inventory,
            "safety_stock": r.safety_stock,
        }
        for r in mrp_rows
        if r.stockout_risk == 1
    ]

    pos_next_two_weeks = [
        {
            "po_number": p.po_number,
            "item_id": p.item_id,
            "supplier": p.supplier,
            "order_qty": p.order_qty,
            "expected_receipt_week": p.expected_receipt_week,
            "status": p.status,
        }
        for p in pos
        if p.expected_receipt_week in ("W1", "W2")
    ]

    # On-time delivery risk: POs arriving after first demand week for that item
    item_demand_weeks = {}
    for r in mrp_rows:
        if r.gross_req > 0:
            idx = week_index(r.week)
            if r.item_id not in item_demand_weeks or idx < item_demand_weeks[r.item_id]:
                item_demand_weeks[r.item_id] = idx

    late_pos = 0
    for p in pos:
        first_demand = item_demand_weeks.get(p.item_id)
        if first_demand is not None:
            receipt_idx = week_index(p.expected_receipt_week)
            if receipt_idx > first_demand:
                late_pos += 1
    on_time_delivery_risk_pct = round((late_pos / len(pos) * 100) if pos else 0, 1)

    # Total planned order value
    total_planned_order_value = 0.0
    for r in mrp_rows:
        unit_cost = item_map[r.item_id].unit_cost if r.item_id in item_map else 0
        total_planned_order_value += r.planned_order * unit_cost
    total_planned_order_value = round(total_planned_order_value, 2)

    # Weeks of supply (avg across items with demand)
    wos_values = []
    item_weeks = {}
    for r in mrp_rows:
        item_weeks.setdefault(r.item_id, []).append(r)
    for item_id, rows in item_weeks.items():
        avg_inv = sum(r.projected_inventory for r in rows) / len(rows)
        demands = [r.gross_req for r in rows if r.gross_req > 0]
        if demands:
            avg_demand = sum(demands) / len(demands)
            if avg_demand > 0:
                wos_values.append(avg_inv / avg_demand)
    weeks_of_supply = round(sum(wos_values) / len(wos_values), 1) if wos_values else 0

    # Critical items: net_req > 0 in W1 or W2
    critical_items = set()
    for r in mrp_rows:
        if r.week in ("W1", "W2") and r.net_req > 0:
            critical_items.add(r.item_id)

    # Demand trend top 3
    item_total_demand = {}
    for r in mrp_rows:
        item_total_demand[r.item_id] = item_total_demand.get(r.item_id, 0) + r.gross_req
    top3 = sorted(item_total_demand.items(), key=lambda x: x[1], reverse=True)[:3]
    top3_ids = [t[0] for t in top3]
    demand_trend = []
    for w in WEEKS:
        row = {"week": w}
        for item_id in top3_ids:
            match = next((r for r in mrp_rows if r.item_id == item_id and r.week == w), None)
            row[item_id] = match.gross_req if match else 0
        demand_trend.append(row)

    # Supply coverage heatmap data
    supply_coverage = []
    for item_id in sorted(item_weeks.keys()):
        row_data = {"item_id": item_id, "weeks": {}}
        for r in item_weeks[item_id]:
            ss = r.safety_stock
            pi = r.projected_inventory
            if pi > ss:
                status = "green"
            elif pi >= ss * 0.8:
                status = "amber"
            else:
                status = "red"
            row_data["weeks"][r.week] = {
                "projected_inventory": pi,
                "safety_stock": ss,
                "status": status,
            }
        supply_coverage.append(row_data)

    return {
        "total_items": len(items),
        "stockout_risks": stockout_risks,
        "pos_next_two_weeks": pos_next_two_weeks,
        "open_pos_count": len(pos),
        "on_time_delivery_risk_pct": on_time_delivery_risk_pct,
        "total_planned_order_value": total_planned_order_value,
        "weeks_of_supply": weeks_of_supply,
        "critical_items_count": len(critical_items),
        "demand_trend_top3": top3_ids,
        "demand_trend": demand_trend,
        "supply_coverage": supply_coverage,
    }


def classify_item_category(bom_df: pd.DataFrame, item_id: str, item_name: str) -> str:
    """Classify item by BOM position, with ID/name fallback."""
    id_upper = str(item_id).strip().upper()
    name_lower = (item_name or "").lower()

    parents = set()
    children = set()
    if bom_df is not None and len(bom_df) > 0:
        parents = set(bom_df["Parent_Item"].astype(str).str.strip().tolist())
        children = set(bom_df["Child_Item"].astype(str).str.strip().tolist())

    is_parent = id_upper in parents
    is_child = id_upper in children

    if is_parent and not is_child:
        return "Finished Good"
    if is_parent and is_child:
        return "Sub-Assembly"
    if is_child and not is_parent:
        if id_upper.startswith("PKG") or "packag" in name_lower:
            return "Packaging"
        return "Raw Material"

    # Not in BOM — fall back to ID prefix / name patterns
    if id_upper.startswith(("ELC", "FG")) or id_upper in {"ITM001", "ITM002"}:
        return "Finished Good"
    if id_upper.startswith("SUB") or id_upper in {"ITM003", "ITM004"}:
        return "Sub-Assembly"
    if id_upper.startswith("PKG") or id_upper in {"ITM008"} or "packag" in name_lower:
        return "Packaging"
    if id_upper.startswith("RAW") or id_upper in {"ITM005", "ITM006", "ITM007"}:
        return "Raw Material"

    if "finished good" in name_lower or "finished goods" in name_lower:
        return "Finished Good"
    if "sub-assembly" in name_lower or "sub assembly" in name_lower:
        return "Sub-Assembly"
    if "packag" in name_lower:
        return "Packaging"
    if "raw material" in name_lower or "raw mat" in name_lower:
        return "Raw Material"

    return "Raw Material"


def category_to_bom_level(category: str) -> int:
    return {
        "Finished Good": 0,
        "Sub-Assembly": 1,
        "Raw Material": 2,
        "Packaging": 2,
    }.get(category, 2)


def build_bom_parents(bom_children: dict) -> dict:
    """Map each child item to list of parent item IDs."""
    bom_parents = {}
    for parent_id, child_list in bom_children.items():
        for child_id, _qty in child_list:
            bom_parents.setdefault(child_id, [])
            if parent_id not in bom_parents[child_id]:
                bom_parents[child_id].append(parent_id)
    return bom_parents


def _parse_plan_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def _parent_production_start_dates(item_id: str, week: str, mrp_data: dict, bom_parents: dict) -> list:
    """Parent production start = parent release date (need date − parent lead time)."""
    dates = []
    for parent_id in bom_parents.get(item_id, []):
        parent_release = mrp_data[parent_id][week].get("release_date")
        parsed = _parse_plan_date(parent_release)
        if parsed:
            dates.append(parsed)
    return dates


def trace_fg_production_date(
    item_id: str,
    week: str,
    mrp_data: dict,
    bom_parents: dict,
    items: dict,
) -> Optional[date]:
    """Walk BOM up to finished goods and return the earliest FG production start date."""
    if items[item_id]["bom_level"] == 0:
        return _parse_plan_date(mrp_data[item_id][week].get("release_date"))

    fg_dates = []
    for parent_id in bom_parents.get(item_id, []):
        if items[parent_id]["bom_level"] == 0:
            parsed = _parse_plan_date(mrp_data[parent_id][week].get("release_date"))
            if parsed:
                fg_dates.append(parsed)
        else:
            traced = trace_fg_production_date(parent_id, week, mrp_data, bom_parents, items)
            if traced:
                fg_dates.append(traced)
    return min(fg_dates) if fg_dates else None


def apply_cascading_release_dates(
    mrp_data: dict,
    items: dict,
    bom_parents: dict,
    planning_start: date,
    today: date,
) -> None:
    """
    Cascade need/release dates through the BOM:
    - Level 0 (FG): need = week start, release = production start
    - Level 1+: need = parent production start, release = need − own lead time
    """
    sorted_ids = sorted(items.keys(), key=lambda x: items[x]["bom_level"])

    for item_id in sorted_ids:
        bom_level = items[item_id]["bom_level"]
        lead_time = items[item_id]["lead_time_weeks"]

        for week in WEEKS:
            data = mrp_data[item_id][week]
            has_activity = (data.get("planned_order") or 0) > 0 or (data.get("gross_req") or 0) > 0
            if not has_activity:
                continue

            if bom_level == 0:
                need_date_obj = week_to_need_date(week, planning_start)
            else:
                parent_starts = _parent_production_start_dates(item_id, week, mrp_data, bom_parents)
                need_date_obj = min(parent_starts) if parent_starts else week_to_need_date(week, planning_start)

            release_date_obj = need_date_obj - timedelta(days=lead_time * 7)
            if bom_level == 0:
                fg_prod = release_date_obj
            else:
                fg_prod = trace_fg_production_date(item_id, week, mrp_data, bom_parents, items)

            data["need_date"] = format_plan_date(need_date_obj)
            data["release_date"] = format_plan_date(release_date_obj)
            data["release_week"] = date_to_week_bucket(release_date_obj, planning_start)
            data["is_overdue"] = 1 if release_date_obj < today else 0
            data["fg_production_date"] = format_plan_date(fg_prod) if fg_prod else None


def topological_sort_items(bom_df: pd.DataFrame, all_items: list) -> list:
    children = set(bom_df["Child_Item"].tolist())
    parents = set(bom_df["Parent_Item"].tolist())
    graph = {}
    for _, row in bom_df.iterrows():
        parent = row["Parent_Item"]
        child = row["Child_Item"]
        graph.setdefault(child, set()).add(parent)

    depth = {}
    for item in all_items:
        depth[item] = 0

    changed = True
    while changed:
        changed = False
        for child in all_items:
            if child in graph:
                max_parent_depth = max(depth.get(p, 0) for p in graph[child])
                new_depth = max_parent_depth + 1
                if new_depth > depth[child]:
                    depth[child] = new_depth
                    changed = True

    return sorted(all_items, key=lambda x: depth[x])


def run_mrp(excel_content: bytes, overrides: Optional[dict] = None) -> dict:
    parsed = parse_excel(excel_content)
    return run_mrp_parsed(parsed, overrides)


def _resolve_available_qty(inv_row) -> float:
    if len(inv_row) == 0:
        return 0.0
    row = inv_row.iloc[0]
    if "On_Hand_Qty" in row.index and "Allocated_Qty" in row.index:
        try:
            return max(0.0, float(row["On_Hand_Qty"]) - float(row["Allocated_Qty"]))
        except (TypeError, ValueError):
            pass
    if "Available_Qty" in row.index:
        try:
            return float(row["Available_Qty"])
        except (TypeError, ValueError):
            pass
    if "On_Hand_Qty" in row.index:
        try:
            return float(row["On_Hand_Qty"])
        except (TypeError, ValueError):
            pass
    return 0.0


def run_mrp_parsed(parsed: dict, overrides: Optional[dict] = None) -> dict:
    item_master = parsed["item_master"]
    bom = parsed["bom"]
    inventory = parsed["inventory"]
    demand_forecast = parsed["demand_forecast"]
    open_pos = parsed["open_pos"]
    sales_history = parsed.get("sales_history")
    if sales_history is None:
        sales_history = pd.DataFrame(columns=["Item_ID"])

    override_item = overrides.get("item_id") if overrides else None
    demand_mult = 1 + (overrides.get("demand_pct", 0) / 100) if overrides else 1
    lead_delta = overrides.get("lead_time_delta", 0) if overrides else 0
    safety_mult = 1 + (overrides.get("safety_stock_pct", 0) / 100) if overrides else 1

    items = {}
    for _, row in item_master.iterrows():
        item_id = str(row["Item_ID"]).strip()
        inv_row = inventory[inventory["Item_ID"] == item_id]
        available = _resolve_available_qty(inv_row)
        safety_stock = float(row["Safety_Stock"])
        lead_time = int(row["Lead_Time_Weeks"])
        if override_item and item_id == override_item:
            safety_stock = max(0, round(safety_stock * safety_mult, 2))
            lead_time = max(0, lead_time + lead_delta)
        category = classify_item_category(bom, item_id, str(row["Item_Name"]))
        items[item_id] = {
            "item_id": item_id,
            "item_name": str(row["Item_Name"]),
            "lead_time_weeks": lead_time,
            "safety_stock": safety_stock,
            "lot_size": float(row["Lot_Size"]),
            "unit": str(row["Unit"]),
            "unit_cost": float(row["Unit_Cost"]),
            "available_qty": available,
            "category": category,
            "bom_level": category_to_bom_level(category),
        }

    bom_children = {}
    for _, row in bom.iterrows():
        parent = str(row["Parent_Item"]).strip()
        child = str(row["Child_Item"]).strip()
        qty = float(row["Qty_Per"])
        bom_children.setdefault(parent, []).append((child, qty))

    demand_weeks = [c for c in demand_forecast.columns if c != "Item_ID"]
    forecast = {}
    for _, row in demand_forecast.iterrows():
        item_id = str(row["Item_ID"]).strip()
        mult = demand_mult if override_item and item_id == override_item else 1
        forecast[item_id] = {
            w: round(float(row[w]) * mult, 2) if w in demand_weeks else 0.0
            for w in WEEKS
        }

    scheduled = {}
    for _, row in open_pos.iterrows():
        item_id = str(row["Item_ID"]).strip()
        week = str(row["Expected_Receipt_Week"]).strip()
        qty = float(row["Order_Qty"])
        scheduled.setdefault(item_id, {}).setdefault(week, 0)
        scheduled[item_id][week] += qty

    all_item_ids = list(items.keys())
    sorted_items = topological_sort_items(bom, all_item_ids)
    planning_start = get_planning_start_monday()
    today = date.today()

    mrp_data = {}
    for item_id in all_item_ids:
        mrp_data[item_id] = {
            w: {
                "gross_req": 0.0,
                "scheduled_receipts": 0.0,
                "projected_inventory": 0.0,
                "net_req": 0.0,
                "planned_order": 0.0,
                "safety_stock": items[item_id]["safety_stock"],
                "stockout_risk": 0,
                "need_date": None,
                "release_date": None,
                "release_week": None,
                "is_overdue": 0,
                "fg_production_date": None,
            }
            for w in WEEKS
        }

    for item_id in sorted_items:
        prev_inv = items[item_id]["available_qty"]
        for week in WEEKS:
            gross = forecast.get(item_id, {}).get(week, 0.0)

            for parent_id, child_list in bom_children.items():
                if any(c == item_id for c, _ in child_list):
                    parent_data = mrp_data[parent_id][week]
                    parent_demand = parent_data["gross_req"] + parent_data["planned_order"]
                    for child, qty_per in child_list:
                        if child == item_id:
                            gross += parent_demand * qty_per

            sched = scheduled.get(item_id, {}).get(week, 0.0)

            net_req = max(0, gross - prev_inv - sched + items[item_id]["safety_stock"])

            lot_size = items[item_id]["lot_size"]
            if net_req > 0:
                planned = math.ceil(net_req / lot_size) * lot_size
            else:
                planned = 0.0

            projected = prev_inv + sched + planned - gross
            stockout = 1 if projected < items[item_id]["safety_stock"] else 0

            mrp_data[item_id][week] = {
                "gross_req": round(gross, 2),
                "scheduled_receipts": round(sched, 2),
                "projected_inventory": round(projected, 2),
                "net_req": round(net_req, 2),
                "planned_order": round(planned, 2),
                "safety_stock": items[item_id]["safety_stock"],
                "stockout_risk": stockout,
                "need_date": None,
                "release_date": None,
                "release_week": None,
                "is_overdue": 0,
                "fg_production_date": None,
            }

            prev_inv = projected

    bom_parents = build_bom_parents(bom_children)
    apply_cascading_release_dates(mrp_data, items, bom_parents, planning_start, today)

    return {
        "items": items,
        "open_pos": open_pos,
        "mrp_data": mrp_data,
        "planning_start": format_plan_date(planning_start),
        "sales_history": sales_history,
        "demand_forecast": demand_forecast,
    }


def _flatten_mrp_result(result: dict) -> list:
    rows = []
    for item_id, weeks in result["mrp_data"].items():
        item = result["items"][item_id]
        for week, data in weeks.items():
            rows.append({
                "item_id": item_id,
                "item_name": item["item_name"],
                "week": week,
                **data,
            })
    return rows


def _scenario_week_row(source: dict) -> dict:
    demand = float(source.get("demand", source.get("gross_req", 0)) or 0)
    stock = float(source.get("stock", source.get("projected_inventory", 0)) or 0)
    planned = float(source.get("planned_order", 0) or 0)
    return {
        "week": source.get("week"),
        "demand": round(demand, 2),
        "stock": round(stock, 2),
        "planned_order": round(planned, 2),
        "gross_req": round(demand, 2),
        "projected_inventory": round(stock, 2),
        "net_req": round(float(source.get("net_req", 0) or 0), 2),
        "stockout_risk": int(source.get("stockout_risk", 0) or 0),
    }


def _scenario_rows_from_db(mrp_rows) -> list:
    by_week = {r.week: r for r in mrp_rows}
    return [
        _scenario_week_row({
            "week": w,
            "gross_req": by_week[w].gross_req if w in by_week else 0,
            "projected_inventory": by_week[w].projected_inventory if w in by_week else 0,
            "planned_order": by_week[w].planned_order if w in by_week else 0,
            "net_req": by_week[w].net_req if w in by_week else 0,
            "stockout_risk": by_week[w].stockout_risk if w in by_week else 0,
        })
        for w in WEEKS
    ]


def _scenario_rows_from_mrp(flat_rows: list, item_id: str) -> list:
    by_week = {r["week"]: r for r in flat_rows if r["item_id"] == item_id}
    return [_scenario_week_row(by_week[w]) if w in by_week else _scenario_week_row({"week": w}) for w in WEEKS]


def _get_source_excel_bytes() -> bytes:
    global last_excel_content
    if last_excel_content:
        return last_excel_content
    filename = sync_state.get("filename") or watch_state.get("watched_file")
    if filename:
        path = os.path.join(WATCHED_DIR, filename)
        if os.path.isfile(path):
            with open(path, "rb") as f:
                return f.read()
    if os.path.isfile(SAMPLE_FILE_PATH):
        with open(SAMPLE_FILE_PATH, "rb") as f:
            return f.read()
    raise HTTPException(status_code=400, detail="No source Excel file available for scenario simulation.")


def clear_and_store(result: dict):
    db = SessionLocal()
    try:
        db.execute(text("DELETE FROM mrp_results"))
        db.execute(text("DELETE FROM items"))
        db.execute(text("DELETE FROM open_pos"))
        db.execute(text("DELETE FROM sales_history"))
        db.execute(text("DELETE FROM forecast_meta"))
        db.commit()

        for item_id, item in result["items"].items():
            db.add(Item(**item))

        for _, row in result["open_pos"].iterrows():
            db.add(
                OpenPO(
                    po_number=str(row["PO_Number"]),
                    item_id=str(row["Item_ID"]),
                    supplier=str(row["Supplier"]),
                    order_qty=float(row["Order_Qty"]),
                    expected_receipt_week=str(row["Expected_Receipt_Week"]),
                    status=str(row["Status"]),
                )
            )

        sales_df = result.get("sales_history")
        if sales_df is not None and len(sales_df) > 0:
            week_cols = [
                c for c in sales_df.columns
                if re.match(r"^W-?\d+$", str(c).strip(), re.IGNORECASE)
            ]
            for _, row in sales_df.iterrows():
                item_id = str(row["Item_ID"]).strip()
                item_name = result["items"].get(item_id, {}).get("item_name", item_id)
                for col in week_cols:
                    label = str(col).strip().upper()
                    try:
                        if label.startswith("W-"):
                            # W-16 … W-1 → offsets -16 … -1 (ascending = oldest→newest)
                            offset = -int(label.split("-")[1])
                        else:
                            # W1 … Wn → offsets 1 … n (W1 oldest, Wn most recent)
                            offset = int(label[1:])
                    except (IndexError, ValueError):
                        continue
                    try:
                        val = float(row[col])
                    except (TypeError, ValueError):
                        val = 0.0
                    db.add(
                        SalesHistory(
                            item_id=item_id,
                            item_name=item_name,
                            week_offset=offset,
                            actual_sales=val,
                            week_label=label if label.startswith("W") else str(col),
                        )
                    )

        for item_id, weeks in result["mrp_data"].items():
            item_name = result["items"][item_id]["item_name"]
            for week, data in weeks.items():
                db.add(
                    MRPResult(
                        item_id=item_id,
                        item_name=item_name,
                        week=week,
                        gross_req=data["gross_req"],
                        scheduled_receipts=data["scheduled_receipts"],
                        projected_inventory=data["projected_inventory"],
                        net_req=data["net_req"],
                        planned_order=data["planned_order"],
                        safety_stock=data["safety_stock"],
                        stockout_risk=data["stockout_risk"],
                        need_date=data["need_date"],
                        release_date=data["release_date"],
                        release_week=data["release_week"],
                        is_overdue=data["is_overdue"],
                        fg_production_date=data.get("fg_production_date"),
                    )
                )

        db.commit()
        create_sql_views()
    finally:
        db.close()


def process_and_store(content: bytes, filename: str, source: str, activate_session: bool = True) -> dict:
    global last_excel_content
    with _process_lock:
        result = run_mrp(content)
        clear_and_store(result)

        last_excel_content = content
        safe_name = os.path.basename(filename or "upload.xlsx")
        dest_path = os.path.join(WATCHED_DIR, safe_name)
        try:
            with open(dest_path, "wb") as f:
                f.write(content)
            # Match watcher mtime so our own write is not re-processed as an external change.
            sync_state["last_mtime"] = os.path.getmtime(dest_path)
        except OSError as exc:
            print(f"Warning: could not persist upload to watched_files: {exc}")

        db = SessionLocal()
        try:
            mrp_rows = db.query(MRPResult).all()
            items = db.query(Item).all()
            pos = db.query(OpenPO).all()
            metrics = compute_summary_metrics(mrp_rows, items, pos)
        finally:
            db.close()

        sync_state["filename"] = filename
        sync_state["last_updated"] = datetime.now(timezone.utc).isoformat()
        sync_state["source"] = source
        watch_state["watched_file"] = filename
        if activate_session:
            session_state["active"] = True

        return {
            "status": "success",
            "message": "MRP calculation complete",
            "items_loaded": metrics["total_items"],
            "pos_found": metrics["open_pos_count"],
            "stockout_risks_detected": len(metrics["stockout_risks"]),
            "filename": filename,
        }


def process_file_from_path(path: str, force: bool = False, activate_session: bool = True) -> Optional[dict]:
    if not path.endswith(".xlsx") or not os.path.isfile(path):
        return None

    mtime = os.path.getmtime(path)
    if not force and mtime <= sync_state["last_mtime"]:
        return None

    try:
        time.sleep(0.3)
        with open(path, "rb") as f:
            content = f.read()
        if len(content) < 64:
            return None
        sync_state["last_mtime"] = mtime
        filename = os.path.basename(path)
        return process_and_store(content, filename, "watch", activate_session=activate_session)
    except Exception as exc:
        print(f"Error processing watched file {path}: {exc}")
        return None


def _schedule_process(path: str):
    global _debounce_timer, _pending_path

    if not path.endswith(".xlsx"):
        return

    def run():
        global _pending_path
        target = _pending_path
        _pending_path = None
        if target:
            process_file_from_path(target)

    with _debounce_lock:
        _pending_path = path
        if _debounce_timer is not None:
            _debounce_timer.cancel()
        _debounce_timer = threading.Timer(1.5, run)
        _debounce_timer.daemon = True
        _debounce_timer.start()


class ExcelWatchHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            _schedule_process(event.src_path)

    def on_modified(self, event):
        if not event.is_directory:
            _schedule_process(event.src_path)


def start_file_watcher():
    global _observer
    handler = ExcelWatchHandler()
    _observer = PollingObserver(timeout=2)
    _observer.schedule(handler, WATCHED_DIR, recursive=False)
    _observer.daemon = True
    _observer.start()
    watch_state["active"] = True

    files = glob.glob(os.path.join(WATCHED_DIR, "*.xlsx"))
    if files:
        newest = max(files, key=os.path.getmtime)
        watch_state["watched_file"] = os.path.basename(newest)
        # Preload DB rows for a fast first render, but do NOT activate the
        # session — the app should start on a clean empty state until the
        # user intentionally uploads or loads sample data.
        process_file_from_path(newest, force=True, activate_session=False)


def stop_file_watcher():
    global _observer
    if _observer is not None:
        _observer.stop()
        _observer.join(timeout=2)
        _observer = None
    watch_state["active"] = False


@app.on_event("startup")
def on_startup():
    start_file_watcher()


@app.on_event("shutdown")
def on_shutdown():
    stop_file_watcher()


@app.post("/validate")
async def validate(file: UploadFile = File(...)):
    content = await file.read()
    return validate_excel(content)


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    content = await file.read()
    try:
        return process_and_store(content, file.filename or "upload.xlsx", "upload")
    except ExcelParseError as exc:
        raise HTTPException(status_code=400, detail={"errors": exc.errors})


@app.get("/session-status")
def get_session_status():
    return {
        "session_active": session_state["active"],
        "filename": sync_state["filename"] if session_state["active"] else None,
        "last_updated": sync_state["last_updated"] if session_state["active"] else None,
    }


@app.post("/load-sample")
def load_sample():
    if not os.path.isfile(SAMPLE_FILE_PATH):
        raise HTTPException(status_code=404, detail="Sample data file not found on the server.")
    with open(SAMPLE_FILE_PATH, "rb") as f:
        content = f.read()
    try:
        return process_and_store(content, "sample_mrp_data.xlsx", "sample")
    except ExcelParseError as exc:
        raise HTTPException(status_code=400, detail={"errors": exc.errors})


@app.post("/clear-data")
def clear_data():
    global last_excel_content
    db = SessionLocal()
    try:
        db.execute(text("DELETE FROM mrp_results"))
        db.execute(text("DELETE FROM items"))
        db.execute(text("DELETE FROM open_pos"))
        db.execute(text("DELETE FROM sales_history"))
        db.execute(text("DELETE FROM forecast_meta"))
        db.commit()
    finally:
        db.close()

    session_state["active"] = False
    sync_state["filename"] = None
    sync_state["last_updated"] = None
    sync_state["source"] = None
    watch_state["watched_file"] = None
    last_excel_content = None
    return {"status": "cleared", "session_active": False}


def _ensure_export_data():
    db = SessionLocal()
    try:
        if db.query(Item).count() == 0:
            raise HTTPException(status_code=400, detail="No MRP data to export. Upload a file first.")
    finally:
        db.close()


def _fetch_export_tables():
    db = SessionLocal()
    try:
        tables = {}
        for name, cols in [
            ("mrp_results", EXPORT_MRP_COLUMNS),
            ("items", EXPORT_ITEMS_COLUMNS),
            ("open_pos", EXPORT_POS_COLUMNS),
        ]:
            result = db.execute(text(f"SELECT {', '.join(cols)} FROM {name}"))
            tables[name] = (cols, [list(row) for row in result.fetchall()])
        return tables
    finally:
        db.close()


def _table_to_csv(columns, rows) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _build_csv_zip(include_instructions: bool = False) -> bytes:
    tables = _fetch_export_tables()
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for table_name, (cols, rows) in tables.items():
            zf.writestr(f"{table_name}.csv", _table_to_csv(cols, rows))
        if include_instructions:
            zf.writestr("instructions.txt", POWERBI_INSTRUCTIONS.encode("utf-8"))
    zip_buf.seek(0)
    return zip_buf.getvalue()


def _auto_width_sheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _write_sheet(ws, columns, rows, title: str):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    _auto_width_sheet(ws)
    ws.title = title[:31]


def _build_excel_workbook() -> bytes:
    tables = _fetch_export_tables()
    db = SessionLocal()
    try:
        mrp_rows = db.query(MRPResult).all()
        items = db.query(Item).all()
        pos = db.query(OpenPO).all()
        metrics = compute_summary_metrics(mrp_rows, items, pos)
    finally:
        db.close()

    wb = Workbook()
    ws_mrp = wb.active
    _write_sheet(ws_mrp, *tables["mrp_results"], "MRP Results")

    ws_items = wb.create_sheet()
    _write_sheet(ws_items, *tables["items"], "Items")

    ws_pos = wb.create_sheet()
    _write_sheet(ws_pos, *tables["open_pos"], "Open POs")

    ws_summary = wb.create_sheet("Summary")
    planned_w1 = sum(r.planned_order for r in mrp_rows if r.week == "W1")
    pos_w1 = sum(1 for p in pos if p.expected_receipt_week == "W1")
    pos_w2 = sum(1 for p in pos if p.expected_receipt_week == "W2")
    summary_rows = [
        ("Metric", "Value"),
        ("Total items", metrics["total_items"]),
        ("Stockout risk count", len(metrics["stockout_risks"])),
        ("Total planned order value", metrics["total_planned_order_value"]),
        ("Planned orders this week (W1)", round(planned_w1, 2)),
        ("Open POs", metrics["open_pos_count"]),
        ("POs due W1", pos_w1),
        ("POs due W2", pos_w2),
        ("Critical items (W1/W2)", metrics["critical_items_count"]),
        ("Weeks of supply (avg)", metrics["weeks_of_supply"]),
    ]
    for row in summary_rows:
        ws_summary.append(list(row))
    ws_summary["A1"].font = Font(bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
    ws_summary["B1"].font = Font(bold=True, color="FFFFFF")
    ws_summary["B1"].fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid")
    _auto_width_sheet(ws_summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.get("/export/powerbi")
def export_powerbi():
    _ensure_export_data()
    content = _build_csv_zip(include_instructions=True)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="mrp_export_powerbi.zip"'},
    )


@app.get("/export/csv")
def export_csv_zip():
    _ensure_export_data()
    content = _build_csv_zip(include_instructions=False)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="mrp_export_csv.zip"'},
    )


@app.get("/export/excel")
def export_excel():
    _ensure_export_data()
    content = _build_excel_workbook()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mrp_export.xlsx"'},
    )


@app.get("/last-updated")
def get_last_updated():
    return {
        "last_updated": sync_state["last_updated"],
        "filename": sync_state["filename"],
    }


@app.get("/metabase-status")
def get_metabase_status():
    try:
        with httpx.Client(timeout=3.0) as client:
            resp = client.get("http://localhost:3000/api/health")
            if resp.status_code == 200:
                data = resp.json()
                return {"running": data.get("status") == "ok"}
    except Exception:
        pass
    return {"running": False}


@app.get("/metabase-config")
def get_metabase_config():
    project_dir = "/Users/nitishsharma/mrp-app"
    if not os.path.isdir(project_dir):
        project_dir = os.path.abspath(os.path.join(BASE_DIR, ".."))
    return {
        "db_path": os.path.abspath(MRP_DB_PATH),
        "metabase_url": "http://localhost:3000",
        "tables": ["mrp_results", "items", "open_pos"],
        "project_dir": project_dir,
        "start_command": f"cd {project_dir} && java -jar metabase.jar",
        "suggested_queries": [
            "SELECT item_name, week, gross_req, net_req FROM mrp_results ORDER BY week",
            "SELECT item_name, week, projected_inventory FROM mrp_results WHERE stockout_risk=1",
            "SELECT supplier, COUNT(*) as po_count, SUM(order_qty) as total_qty FROM open_pos GROUP BY supplier",
        ],
    }


@app.get("/watch-status")
def get_watch_status():
    watched = watch_state["watched_file"]
    if not watched:
        files = glob.glob(os.path.join(WATCHED_DIR, "*.xlsx"))
        if files:
            watched = os.path.basename(max(files, key=os.path.getmtime))
    return {
        "active": watch_state["active"],
        "filename": watched,
        "watched_folder_path": WATCHED_DIR,
    }


@app.get("/sync-status")
def get_sync_status():
    db = SessionLocal()
    try:
        has_data = db.query(Item).count() > 0
    finally:
        db.close()

    return {
        "active": watch_state["active"],
        "has_data": has_data,
        "filename": sync_state["filename"],
        "last_updated": sync_state["last_updated"],
        "watched_folder_path": WATCHED_DIR,
    }


@app.post("/sync/open-folder")
def open_watched_folder():
    try:
        if os.name == "posix":
            subprocess.Popen(["open", WATCHED_DIR])
        return {"status": "ok", "path": WATCHED_DIR}
    except Exception as e:
        return {"status": "error", "message": str(e), "path": WATCHED_DIR}


@app.get("/items")
def get_items():
    db = SessionLocal()
    try:
        rows = db.query(Item).all()
        return [
            {
                "item_id": r.item_id,
                "item_name": r.item_name,
                "lead_time_weeks": r.lead_time_weeks,
                "safety_stock": r.safety_stock,
                "lot_size": r.lot_size,
                "unit": r.unit,
                "unit_cost": r.unit_cost,
                "available_qty": r.available_qty,
                "category": r.category,
                "bom_level": r.bom_level,
            }
            for r in rows
        ]
    finally:
        db.close()


def _sales_series_for_item(db, item_id: str) -> Tuple[list, list]:
    """Return (week_labels oldest→newest, actual values) for an item.

    W1..Wn: W1 oldest, Wn most recent (immediately before forecast).
    W-N..W-1: more-negative offsets are older.
    """
    rows = (
        db.query(SalesHistory)
        .filter(SalesHistory.item_id == item_id)
        .order_by(SalesHistory.week_offset.asc())
        .all()
    )
    labels = [r.week_label for r in rows]
    values = [float(r.actual_sales or 0) for r in rows]
    return labels, values


@app.get("/sales-history")
def get_sales_history(item_id: Optional[str] = None):
    db = SessionLocal()
    try:
        q = db.query(SalesHistory)
        if item_id:
            q = q.filter(SalesHistory.item_id == item_id)
        rows = q.order_by(SalesHistory.item_id, SalesHistory.week_offset.asc()).all()
        return [
            {
                "item_id": r.item_id,
                "item_name": r.item_name,
                "week_offset": r.week_offset,
                "week_label": r.week_label,
                "actual_sales": r.actual_sales,
            }
            for r in rows
        ]
    finally:
        db.close()


@app.get("/forecast/status")
def get_forecast_status():
    db = SessionLocal()
    try:
        has_history = db.query(SalesHistory).count() > 0
        active = (
            db.query(ForecastMeta)
            .filter(ForecastMeta.active == 1)
            .order_by(ForecastMeta.id.desc())
            .first()
        )
        return {
            "has_sales_history": has_history,
            "forecast_active": bool(active),
            "method": active.method if active else None,
            "item_id": active.item_id if active else None,
            "applied_at": active.applied_at if active else None,
        }
    finally:
        db.close()


@app.post("/forecast")
def run_forecast(req: ForecastRequest):
    method = req.method
    allowed = {"moving_average", "exponential_smoothing", "double_exponential", "holt_winters"}
    if method not in allowed:
        raise HTTPException(status_code=400, detail=f"method must be one of {sorted(allowed)}")

    db = SessionLocal()
    try:
        item = db.query(Item).filter(Item.item_id == req.item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"Item {req.item_id} not found")

        labels, values = _sales_series_for_item(db, req.item_id)
        if len(values) < 4:
            raise HTTPException(
                status_code=400,
                detail="Need at least 4 weeks of sales history to generate a forecast. "
                "Add a Sales_History sheet to your Excel upload.",
            )

        params = (req.params.dict() if req.params else {})
        periods = max(1, min(int(req.periods or 8), 8))
        try:
            fc_values = forecast_with_method(values, method, periods=periods, params=params)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Forecast failed: {exc}") from exc

        import numpy as np

        lower, upper = confidence_bounds(fc_values, np.asarray(values, dtype=float))
        accuracy = holdout_accuracy(values, method, params=params, holdout=min(4, len(values) // 3 or 1))

        # Persist draft forecast meta (not applied yet)
        draft = ForecastMeta(
            item_id=req.item_id,
            method=method,
            applied_at=None,
            mape=accuracy["mape"],
            mad=accuracy["mad"],
            bias=accuracy["bias"],
            forecast_json=json.dumps(fc_values),
            overrides_json=json.dumps({}),
            active=0,
        )
        db.add(draft)
        db.commit()

        historical = [{"week": labels[i], "actual": values[i]} for i in range(len(values))]
        forecast = [
            {
                "week": WEEKS[i] if i < len(WEEKS) else f"W{i + 1}",
                "forecast": fc_values[i],
                "lower": lower[i],
                "upper": upper[i],
            }
            for i in range(len(fc_values))
        ]

        return {
            "item_id": req.item_id,
            "item_name": item.item_name,
            "method": method,
            "historical": historical,
            "forecast": forecast,
            "accuracy": accuracy,
        }
    finally:
        db.close()


@app.post("/forecast/apply")
def apply_forecast(req: ForecastApplyRequest):
    global last_excel_content, sync_state

    if not req.forecast_values or len(req.forecast_values) < 1:
        raise HTTPException(status_code=400, detail="forecast_values required")

    values = [max(0.0, float(v)) for v in req.forecast_values]
    while len(values) < 8:
        values.append(values[-1] if values else 0.0)
    values = values[:8]

    content = last_excel_content
    if not content:
        try:
            content = _get_source_excel_bytes()
        except HTTPException:
            raise HTTPException(status_code=400, detail="No source Excel available to re-run MRP")

    parsed = parse_excel(content)
    demand = parsed["demand_forecast"].copy()
    item_id = str(req.item_id).strip()

    if "Item_ID" not in demand.columns:
        raise HTTPException(status_code=400, detail="Demand forecast sheet missing Item_ID")

    mask = demand["Item_ID"].astype(str).str.strip() == item_id
    if not mask.any():
        # Insert new demand row for this finished good
        new_row = {"Item_ID": item_id}
        for i, w in enumerate(WEEKS):
            new_row[w] = values[i]
        demand = pd.concat([demand, pd.DataFrame([new_row])], ignore_index=True)
    else:
        for i, w in enumerate(WEEKS):
            if w not in demand.columns:
                demand[w] = 0.0
            demand.loc[mask, w] = values[i]

    parsed["demand_forecast"] = demand
    result = run_mrp_parsed(parsed)
    clear_and_store(result)
    last_excel_content = content

    applied_at = datetime.now(timezone.utc).isoformat()
    db = SessionLocal()
    try:
        db.query(ForecastMeta).update({ForecastMeta.active: 0})
        meta = ForecastMeta(
            item_id=item_id,
            method=req.method or "applied",
            applied_at=applied_at,
            mape=None,
            mad=None,
            bias=None,
            forecast_json=json.dumps(values),
            overrides_json=json.dumps({}),
            active=1,
        )
        db.add(meta)
        db.commit()

        mrp_rows = db.query(MRPResult).filter(MRPResult.item_id == item_id).all()
        sync_state["last_updated"] = applied_at
        sync_state["filename"] = sync_state.get("filename") or "forecast_applied.xlsx"

        return {
            "status": "success",
            "message": "Forecast applied — MRP recalculated with new demand plan",
            "item_id": item_id,
            "method": req.method,
            "applied_at": applied_at,
            "demand": {WEEKS[i]: values[i] for i in range(8)},
            "mrp_results": [
                {
                    "week": r.week,
                    "gross_req": r.gross_req,
                    "planned_order": r.planned_order,
                    "projected_inventory": r.projected_inventory,
                }
                for r in mrp_rows
            ],
        }
    finally:
        db.close()


@app.post("/forecast/override")
def override_forecast(req: ForecastOverrideRequest):
    db = SessionLocal()
    try:
        meta = (
            db.query(ForecastMeta)
            .filter(ForecastMeta.item_id == req.item_id)
            .order_by(ForecastMeta.id.desc())
            .first()
        )
        if not meta:
            raise HTTPException(status_code=404, detail="No forecast found for this item. Generate one first.")
        overrides = json.loads(meta.overrides_json or "{}")
        overrides[req.week] = float(req.value)
        meta.overrides_json = json.dumps(overrides)
        db.commit()
        return {"status": "ok", "item_id": req.item_id, "week": req.week, "value": req.value, "overrides": overrides}
    finally:
        db.close()


@app.get("/forecast/accuracy")
def get_forecast_accuracy():
    db = SessionLocal()
    try:
        metas = db.query(ForecastMeta).order_by(ForecastMeta.id.desc()).all()
        seen = set()
        out = []
        for m in metas:
            if m.item_id in seen:
                continue
            seen.add(m.item_id)
            out.append({
                "item_id": m.item_id,
                "method": m.method,
                "mape": m.mape,
                "mad": m.mad,
                "bias": m.bias,
                "active": bool(m.active),
                "applied_at": m.applied_at,
            })
        return out
    finally:
        db.close()


@app.get("/mrp-results")
def get_mrp_results():
    db = SessionLocal()
    try:
        rows = (
            db.query(MRPResult)
            .order_by(MRPResult.item_id, MRPResult.week)
            .all()
        )
        return [
            {
                "item_id": r.item_id,
                "item_name": r.item_name,
                "week": r.week,
                "gross_req": r.gross_req,
                "scheduled_receipts": r.scheduled_receipts,
                "projected_inventory": r.projected_inventory,
                "net_req": r.net_req,
                "planned_order": r.planned_order,
                "safety_stock": r.safety_stock,
                "stockout_risk": r.stockout_risk,
                "need_date": r.need_date,
                "release_date": r.release_date,
                "release_week": r.release_week,
                "is_overdue": r.is_overdue,
                "fg_production_date": r.fg_production_date,
            }
            for r in rows
        ]
    finally:
        db.close()


@app.get("/open-pos")
def get_open_pos():
    db = SessionLocal()
    try:
        rows = db.query(OpenPO).all()
        return [
            {
                "po_number": r.po_number,
                "item_id": r.item_id,
                "supplier": r.supplier,
                "order_qty": r.order_qty,
                "expected_receipt_week": r.expected_receipt_week,
                "status": r.status,
            }
            for r in rows
        ]
    finally:
        db.close()


@app.get("/summary")
def get_summary():
    db = SessionLocal()
    try:
        mrp_rows = db.query(MRPResult).all()
        items = db.query(Item).all()
        pos = db.query(OpenPO).all()

        if not items:
            return {
                "total_items": 0,
                "stockout_risks": [],
                "pos_next_two_weeks": [],
                "open_pos_count": 0,
                "on_time_delivery_risk_pct": 0,
                "total_planned_order_value": 0,
                "weeks_of_supply": 0,
                "critical_items_count": 0,
                "demand_trend_top3": [],
                "demand_trend": [],
                "supply_coverage": [],
            }

        metrics = compute_summary_metrics(mrp_rows, items, pos)
        metrics["loaded_at"] = sync_state.get("last_updated")
        return metrics
    finally:
        db.close()


@app.get("/exceptions")
def get_exceptions():
    db = SessionLocal()
    try:
        mrp_rows = db.query(MRPResult).order_by(MRPResult.item_id, MRPResult.week).all()
        items = db.query(Item).all()
        pos = db.query(OpenPO).all()

        if not mrp_rows:
            return {
                "health_score": 100,
                "health_color": "green",
                "overdue_releases": [],
                "critical_stock_alerts": [],
                "pos_arriving_this_week": [],
                "no_coverage": [],
            }

        today = date.today()
        item_map = {i.item_id: i for i in items}

        all_item_ids = {r.item_id for r in mrp_rows}
        at_risk_items = {r.item_id for r in mrp_rows if r.stockout_risk == 1}
        safe_count = len(all_item_ids - at_risk_items)
        health_score = round((safe_count / len(all_item_ids)) * 100) if all_item_ids else 100
        if health_score > 80:
            health_color = "green"
        elif health_score >= 60:
            health_color = "amber"
        else:
            health_color = "red"

        overdue_releases = []
        for r in mrp_rows:
            if r.planned_order <= 0 or not r.release_date:
                continue
            try:
                release_dt = datetime.strptime(r.release_date, "%Y-%m-%d").date()
            except ValueError:
                continue
            if release_dt >= today:
                continue
            days_overdue = (today - release_dt).days
            overdue_releases.append({
                "item_id": r.item_id,
                "item_name": r.item_name,
                "planned_qty": r.planned_order,
                "was_due": r.release_date,
                "need_by": r.need_date,
                "days_overdue": days_overdue,
            })
        overdue_releases.sort(key=lambda x: -x["days_overdue"])

        critical_stock_alerts = []
        for r in mrp_rows:
            if r.week not in ("W1", "W2"):
                continue
            if r.projected_inventory >= r.safety_stock:
                continue
            shortage = round(r.safety_stock - r.projected_inventory, 2)
            action = "Release PO immediately" if r.planned_order > 0 else "Create planned order"
            critical_stock_alerts.append({
                "item_id": r.item_id,
                "item_name": r.item_name,
                "week": r.week,
                "current_stock": r.projected_inventory,
                "safety_stock": r.safety_stock,
                "shortage": shortage,
                "action_needed": action,
            })

        pos_arriving_this_week = [
            {
                "po_number": p.po_number,
                "item_id": p.item_id,
                "item_name": item_map[p.item_id].item_name if p.item_id in item_map else p.item_id,
                "supplier": p.supplier,
                "qty": p.order_qty,
                "status": p.status,
            }
            for p in pos
            if p.expected_receipt_week == "W1"
        ]

        no_coverage = []
        for r in mrp_rows:
            if r.net_req <= 0:
                continue
            if r.scheduled_receipts > 0 or r.planned_order > 0:
                continue
            no_coverage.append({
                "item_id": r.item_id,
                "item_name": r.item_name,
                "week": r.week,
                "net_req": r.net_req,
                "projected_inventory": r.projected_inventory,
            })

        return {
            "health_score": health_score,
            "health_color": health_color,
            "overdue_releases": overdue_releases,
            "critical_stock_alerts": critical_stock_alerts,
            "pos_arriving_this_week": pos_arriving_this_week,
            "no_coverage": no_coverage,
        }
    finally:
        db.close()


@app.post("/scenario")
def run_scenario(req: ScenarioRequest):
    if not session_state["active"]:
        raise HTTPException(status_code=400, detail="No active session. Upload data first.")

    item_id = (req.item_id or "").strip()
    demand_change = req.demand_pct_change if req.demand_pct_change is not None else req.demand_pct
    lead_delay = req.lead_time_delay_weeks if req.lead_time_delay_weeks is not None else req.lead_time_delta
    safety_change = req.safety_stock_pct_change if req.safety_stock_pct_change is not None else req.safety_stock_pct
    demand_change = max(-50, min(50, demand_change))
    lead_delay = max(0, min(4, lead_delay))
    safety_change = max(-50, min(50, safety_change))

    print(
        f"Scenario request: item_id={item_id}, demand_change={demand_change}, "
        f"lead_delay={lead_delay}, safety_change={safety_change}"
    )

    db = SessionLocal()
    try:
        item = db.query(Item).filter(Item.item_id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"Item {item_id} not found in current plan.")
        current_db_rows = (
            db.query(MRPResult)
            .filter(MRPResult.item_id == item_id)
            .order_by(MRPResult.week)
            .all()
        )
    finally:
        db.close()

    print(f"Current MRP rows found: {len(current_db_rows)}")

    content = _get_source_excel_bytes()
    overrides = {
        "item_id": item_id,
        "demand_pct": demand_change,
        "lead_time_delta": lead_delay,
        "safety_stock_pct": safety_change,
    }
    scenario_result = run_mrp(content, overrides=overrides)
    scenario_flat = _flatten_mrp_result(scenario_result)
    scenario_rows = [r for r in scenario_flat if r["item_id"] == item_id]

    print(f"Scenario result rows: {len(scenario_rows)}")

    current = _scenario_rows_from_db(current_db_rows)
    scenario = _scenario_rows_from_mrp(scenario_flat, item_id)

    current_by_week = {r["week"]: r for r in current}
    scenario_by_week = {r["week"]: r for r in scenario}

    stockout_before = sum(1 for r in current if r["stockout_risk"] == 1)
    stockout_after = sum(1 for r in scenario if r["stockout_risk"] == 1)
    stockout_change = stockout_after - stockout_before

    additional_orders = 0
    additional_cost = 0.0
    for w in WEEKS:
        c_row = current_by_week[w]
        s_row = scenario_by_week[w]
        delta = max(0, s_row["planned_order"] - c_row["planned_order"])
        if delta > 0:
            additional_orders += 1
            additional_cost += delta * float(item.unit_cost or 0)

    return {
        "item_id": item_id,
        "current": current,
        "scenario": scenario,
        "impact": {
            "stockout_risk_before": stockout_before,
            "stockout_risk_after": stockout_after,
            "stockout_risk_change": stockout_change,
            "additional_planned_orders": additional_orders,
            "additional_orders": additional_orders,
            "additional_cost": round(additional_cost, 2),
        },
    }


@app.get("/pipeline-status")
def get_pipeline_status():
    db = SessionLocal()
    try:
        item_count = db.query(Item).count()
        mrp_count = db.query(MRPResult).count()
        po_count = db.query(OpenPO).count()
    finally:
        db.close()

    has_excel = bool(sync_state.get("filename") or watch_state.get("watched_file"))
    has_database = item_count > 0
    has_ai = bool(ANTHROPIC_API_KEY)

    return {
        "steps": [
            {
                "id": "excel",
                "label": "Excel File",
                "description": sync_state.get("filename") or watch_state.get("watched_file") or "Waiting for upload",
                "active": has_excel,
            },
            {
                "id": "engine",
                "label": "Python MRP Engine",
                "description": "FastAPI + pandas BOM explosion",
                "active": True,
            },
            {
                "id": "database",
                "label": "SQL Database",
                "description": f"{DB_TYPE.upper()} · {item_count} items, {mrp_count} MRP rows, {po_count} POs",
                "active": has_database,
            },
            {
                "id": "ai",
                "label": "AI Layer",
                "description": "Claude reads live SQL data for chat",
                "active": has_ai and has_database,
            },
            {
                "id": "dashboard",
                "label": "Dashboard",
                "description": "React UI powered by REST + SQL views",
                "active": has_database,
            },
        ],
        "database_type": DB_TYPE,
        "database_display": get_db_display_url(),
        "postgres_fallback_reason": PG_FALLBACK_REASON,
    }


@app.get("/database/info")
def get_database_info():
    return {
        "type": DB_TYPE,
        "display_url": get_db_display_url(),
        "postgres_fallback_reason": PG_FALLBACK_REASON,
        "tables": ["items", "mrp_results", "open_pos"],
        "views": list(SQL_VIEWS.keys()),
    }


@app.get("/database/tables")
def get_database_tables():
    tables = ["items", "mrp_results", "open_pos"]
    db = SessionLocal()
    try:
        counts = {}
        for table in tables:
            result = db.execute(text(f"SELECT COUNT(*) FROM {table}"))
            counts[table] = result.scalar() or 0
        return {"tables": [{"name": t, "row_count": counts[t]} for t in tables]}
    finally:
        db.close()


@app.post("/database/query")
def run_sql_query_legacy(req: SqlQueryRequest):
    return execute_sql_query(req.query)


@app.get("/sql-query")
def sql_query_get(query: str = Query(..., description="SQL SELECT query")):
    return execute_sql_query(query)


@app.post("/sql-query")
def sql_query_post(req: SqlQueryRequest = Body(...)):
    return execute_sql_query(req.query)


def execute_sql_query(sql: str):
    sql = (sql or "").strip()
    if not sql:
        return {"error": "Please enter a query.", "columns": [], "rows": [], "count": 0}
    if not is_readonly_query(sql):
        return {"error": "Only read-only SELECT, WITH, or PRAGMA queries are allowed.", "columns": [], "rows": [], "count": 0}

    db = SessionLocal()
    try:
        result = db.execute(text(sql))
        if not result.returns_rows:
            return {"error": "Query did not return rows.", "columns": [], "rows": [], "count": 0}
        columns = list(result.keys())
        rows = []
        for i, row in enumerate(result):
            if i >= 500:
                break
            mapping = row._mapping
            rows.append([_serialize_cell(mapping[col]) for col in columns])
        return {"columns": columns, "rows": rows, "count": len(rows), "truncated": len(rows) >= 500}
    except Exception as exc:
        return {"error": str(exc), "columns": [], "rows": [], "count": 0}
    finally:
        db.close()


def _serialize_cell(value):
    if value is None:
        return None
    if isinstance(value, float):
        return round(value, 4) if value != int(value) else int(value)
    return value


@app.get("/database/views")
def get_database_views():
    db = SessionLocal()
    try:
        views = []
        for name in SQL_VIEWS:
            create_sql = VIEW_CREATE_SQL[name]
            result = db.execute(text(f"SELECT * FROM {name} LIMIT 100"))
            columns = list(result.keys())
            rows = [
                [_serialize_cell(row._mapping[col]) for col in columns]
                for row in result
            ]
            views.append({
                "name": name,
                "create_sql": create_sql.strip(),
                "columns": columns,
                "rows": rows,
                "row_count": len(rows),
            })
        return {"views": views}
    except Exception as exc:
        return {"views": [], "error": str(exc)}
    finally:
        db.close()


def _build_data_context(items, pos, mrp_rows) -> str:
    data_context = "ITEMS:\n"
    for item in items:
        data_context += (
            f"  {item.item_id} ({item.item_name}): lead_time={item.lead_time_weeks}w, "
            f"safety_stock={item.safety_stock}, lot_size={item.lot_size}, "
            f"available={item.available_qty}, unit_cost=${item.unit_cost}\n"
        )

    data_context += "\nOPEN PURCHASE ORDERS:\n"
    for p in pos:
        data_context += (
            f"  {p.po_number}: {p.item_id} qty={p.order_qty} week={p.expected_receipt_week} "
            f"supplier={p.supplier} status={p.status}\n"
        )

    data_context += "\nMRP RESULTS:\n"
    for r in mrp_rows:
        risk = "RISK" if r.stockout_risk else "OK"
        data_context += (
            f"  {r.item_id} {r.week}: gross={r.gross_req}, sched={r.scheduled_receipts}, "
            f"proj_inv={r.projected_inventory}, net={r.net_req}, planned={r.planned_order}, "
            f"safety={r.safety_stock}, {risk}\n"
        )
    return data_context


def _is_w1_po_cost_chart_query(question: str) -> bool:
    q = question.lower()
    po_terms = any(kw in q for kw in (
        "po cost", "po costs", "purchase order", "purchase orders",
        "open po", "open pos", "po arrival", "pos arriving",
    ))
    week_terms = any(kw in q for kw in ("this week", "w1", "week 1", "arriving this week"))
    return po_terms and week_terms


def _build_w1_po_cost_chart(items, pos) -> Optional[dict]:
    item_map = {i.item_id: i for i in items}
    data = []
    for p in sorted(pos, key=lambda x: x.po_number):
        if p.expected_receipt_week != "W1":
            continue
        item = item_map.get(p.item_id)
        unit_cost = float(item.unit_cost) if item else 0.0
        data.append({
            "name": p.po_number,
            "value": round(float(p.order_qty) * unit_cost, 2),
        })
    if not data:
        return None
    return {
        "type": "bar",
        "title": "W1 PO Arrival Costs",
        "subtitle": "Total cost of purchase orders arriving this week",
        "data": data,
        "xKey": "name",
        "yKeys": ["value"],
        "colors": [],
    }


async def _call_claude(client: httpx.AsyncClient, api_key: str, system: Optional[str], user_content: str) -> tuple[bool, str]:
    payload = {
        "model": "claude-sonnet-4-6",
        "max_tokens": 4096,
        "messages": [{"role": "user", "content": user_content}],
    }
    if system:
        payload["system"] = system

    response = await client.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json=payload,
    )
    if response.status_code != 200:
        return False, f"Claude API error: {response.status_code} - {response.text}"
    return True, response.json()["content"][0]["text"]


@app.post("/chat")
async def chat(req: ChatRequest):
    db = SessionLocal()
    try:
        mrp_rows = db.query(MRPResult).order_by(MRPResult.item_id, MRPResult.week).all()
        items = db.query(Item).all()
        pos = db.query(OpenPO).all()

        if not mrp_rows:
            return {
                "reply": "No MRP data available. Please upload an Excel file first using the Upload tab.",
                "chart": None,
            }

        data_context = _build_data_context(items, pos, mrp_rows)
        mrp_data = [
            {
                "item_id": r.item_id,
                "week": r.week,
                "gross_req": r.gross_req,
                "scheduled_receipts": r.scheduled_receipts,
                "projected_inventory": r.projected_inventory,
                "net_req": r.net_req,
                "planned_order": r.planned_order,
                "safety_stock": r.safety_stock,
                "stockout_risk": r.stockout_risk,
            }
            for r in mrp_rows
        ]
        api_key = req.api_key or ANTHROPIC_API_KEY

        text_system = (
            "You are an MRP (Material Requirements Planning) analyst assistant. "
            "Answer the user's question clearly and concisely using the data provided. "
            "Use markdown formatting where helpful.\n\n"
            f"DATA CONTEXT:\n{data_context}"
        )

        user_question = req.message
        chart_keywords = ["chart", "graph", "plot", "visualize", "show me", "trend", "bar", "line", "pie"]
        wants_chart = any(kw in user_question.lower() for kw in chart_keywords)

        chart_system = (
            "You are a data formatter. Respond with ONLY a JSON object, no other text whatsoever.\n"
            'Format: {"type":"bar","title":"string","subtitle":"string","data":[{"name":"string","value":number}]}\n'
            "Do not include any explanation, markdown, or text outside the JSON object."
        )
        chart_user_message = (
            f"Generate chart data for this request: '{user_question}'\n\n"
            "Use this MRP data to build the chart values:\n"
            f"{json.dumps(mrp_data[:200])}\n\n"
            "Return ONLY the JSON object."
        )

        async with httpx.AsyncClient(timeout=60.0) as client:
            # First call: the text answer.
            ok_text, text_result = await _call_claude(client, api_key, text_system, user_question)
            if not ok_text:
                return {"reply": text_result, "chart": None}

            # Second call: chart data only (only when the request looks chart-related).
            chart = None
            if wants_chart:
                if _is_w1_po_cost_chart_query(user_question):
                    chart = _build_w1_po_cost_chart(items, pos)
                    print(f"[chat] built deterministic W1 PO cost chart: {chart}")
                else:
                    ok_chart, chart_result = await _call_claude(client, api_key, chart_system, chart_user_message)
                    if ok_chart:
                        print(f"[chat] raw chart response: {chart_result}")
                        chart = _parse_chart_only(chart_result)
                    else:
                        print(f"[chat] chart call failed: {chart_result}")

        return {"reply": text_result, "chart": chart}
    finally:
        db.close()


def _parse_chart_only(text: str) -> Optional[dict]:
    raw = (text or "").strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    if raw.lower() == "null" or raw == "":
        return None
    try:
        chart_data = json.loads(raw)
    except json.JSONDecodeError:
        parsed = _coerce_json_dict(raw)
        if parsed is None:
            print(f"[chat] could not parse chart JSON: {raw[:200]}")
            return None
        chart_data = parsed
    return _normalize_chart(chart_data)


def _normalize_chart(chart) -> Optional[dict]:
    if not isinstance(chart, dict):
        return None
    chart_type = chart.get("type")
    data = chart.get("data")
    if chart_type not in ("bar", "line", "pie", "heatmap") or not isinstance(data, list) or len(data) == 0:
        return None
    return {
        "type": chart_type,
        "title": chart.get("title") or "",
        "subtitle": chart.get("subtitle") or "",
        "data": data,
        "xKey": chart.get("xKey") or ("week" if chart_type == "line" else "name"),
        "yKeys": chart.get("yKeys") or ["value"],
        "colors": chart.get("colors") or [],
    }


def _coerce_json_dict(raw: str) -> Optional[dict]:
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass
    # Fallback: extract the first balanced {...} block from the text.
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            parsed = json.loads(raw[start:end + 1])
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
    return None
